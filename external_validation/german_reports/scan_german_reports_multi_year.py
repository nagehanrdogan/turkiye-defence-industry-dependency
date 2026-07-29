"""
scan_german_reports_multi_year.py
-----------------------------------
Almanya'nin yillik Ruestungsexportbericht'lerini (2014-2024, dogrulanmis 9 yil)
otomatik indirir ve her birinde 'Turkei' gecen yerleri tarar -- ABD/DSCA
verisiyle AYNI ZAMAN ARALIGINDA (coklu yil) bir karsilastirma yapabilmeniz icin.

NASIL BULUNDU: Her yilin PDF linki, o yilin .html tanitim sayfasindaki
"Download (PDF, X MB)" linkinden tek tek dogrulandi (Claude tarafindan,
web_fetch araciyla). Site her yil icin farkli bir isimlendirme kullaniyor
(2022/2023 diger yillardan farkli), bu yuzden tahmin degil, dogrulanmis
linkler kullanildi.

EKSIK YILLAR: 2013 ve oncesi (2009-2013) icin dogrudan link bulunamadi --
bu donem farkli bir yayinlama rejimine sahipti ("Zwischen 1999 und 2012
wurden... Jahresberichte veroeffentlicht"). Bu yillari tamamlamak isterseniz:
https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2013.html
adresine gidip "Download (PDF...)" linkini bulun, asagidaki REPORTS
sozlugune elle ekleyin. 2009-2012 icin ise:
http://ruestungsexport-info.de/zahlen-fakten/ruestungsexportberichte-der-bundesregierung.html
sayfasi TUM yillara (1999'a kadar) link veriyor -- oradan tamamlayabilirsiniz.

Kullanim:
    pip3 install requests pdfplumber openpyxl --break-system-packages
    python3 scan_german_reports_multi_year.py --output almanya_coklu_yil.xlsx
"""

import argparse
import os
import re
import time

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )
}

# Dogrulanmis, calisan PDF linkleri (yil: url)
REPORTS = {
    2024: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2024.pdf?__blob=publicationFile&v=1",
    2023: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/bericht-der-bundesregierung-ueber-ihre-exportpolitik-fuer-konventionelle-ruestungsgueter-im-jahre-2023.pdf?__blob=publicationFile&v=1",
    2022: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Downloads/B/bericht-bundesregierung-exportpolitik-konventionelle-ruestungsgueter-2022.pdf?__blob=publicationFile&v=1",
    2021: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2021.pdf?__blob=publicationFile&v=1",
    2018: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2018.pdf?__blob=publicationFile&v=1",
    2017: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2017.pdf?__blob=publicationFile&v=1",
    2016: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2016.pdf?__blob=publicationFile",
    2015: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2015.pdf?__blob=publicationFile&v=1",
    2014: "https://www.bundeswirtschaftsministerium.de/Redaktion/DE/Publikationen/Aussenwirtschaft/ruestungsexportbericht-2014.pdf?__blob=publicationFile&v=1",
    # 2013 ve oncesi icin bkz. yukaridaki not -- elle ekleyin:
    # 2013: "https://.../ruestungsexportbericht-2013.pdf?__blob=publicationFile&v=X",
}


def download_pdf(year: int, url: str, download_dir: str = "german_reports") -> str:
    os.makedirs(download_dir, exist_ok=True)
    local_path = os.path.join(download_dir, f"ruestungsexportbericht_{year}.pdf")
    if os.path.exists(local_path):
        return local_path
    print(f"  {year} indiriliyor...")
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    with open(local_path, "wb") as f:
        f.write(resp.content)
    time.sleep(1)
    return local_path


def search_pdf_for_turkey(pdf_path: str, keyword: str = "Türkei") -> list[str]:
    import pdfplumber

    hits = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if keyword in text:
                idx = text.find(keyword)
                snippet = text[max(0, idx - 150): idx + 250]
                hits.append({"page": i + 1, "snippet": snippet})
    return hits


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="almanya_coklu_yil.xlsx")
    args = parser.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = "Turkei_Gecen_Yerler"
    headers = ["Yil", "Sayfa", "Metin_Parcasi"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    row = 2
    for year, url in sorted(REPORTS.items()):
        print(f"=== {year} ===")
        try:
            local_path = download_pdf(year, url)
            hits = search_pdf_for_turkey(local_path)
            print(f"  -> {len(hits)} sayfada 'Türkei' bulundu.")
            for h in hits:
                ws.cell(row=row, column=1, value=year)
                ws.cell(row=row, column=2, value=h["page"])
                ws.cell(row=row, column=3, value=h["snippet"].replace("\n", " "))
                row += 1
        except Exception as e:
            print(f"  [HATA] {year}: {e}")
            ws.cell(row=row, column=1, value=year)
            ws.cell(row=row, column=3, value=f"HATA: {e}")
            row += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 120
    ws.freeze_panes = "A2"

    wb.save(args.output)
    print(f"\nYazildi: {args.output}")


if __name__ == "__main__":
    main()
