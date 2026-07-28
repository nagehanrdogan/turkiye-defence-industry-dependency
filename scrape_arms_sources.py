"""
scrape_arms_sources.py
-----------------------
Turkiye'nin silah tedarik kaynaklarini otomatik toplayan scraper.

ONEMLI - BUNU CALISTIRMADAN ONCE OKUYUN:
- Bu script SIZIN bilgisayarinizda calisacak sekilde tasarlandi (Claude'un
  sanal makinesi bu sitelere internet erisimi olmadigi icin buradan test
  edilemedi). Ilk calistirmada bazi selector/regex'lerin site yapisina tam
  oturmama ihtimali var -- --debug bayragiyla calistirirsaniz script, bulamadigi
  sayfalarda ham HTML'in bir kismini ekrana basar, boylece neyin degistigini
  gorup regex'i güncelleyebilirsiniz.
- Siteler bot trafigini engelleyebilir; User-Agent header'i bu yuzden eklendi.
  Yine de calismazsa, tarayicinizin "Sayfa Kaynagini Goruntule" (View Source)
  ozelligiyle gercek HTML yapisini kontrol edip BeautifulSoup sorgularini
  (soup.find_all(...)) ona gore duzenleyin.
- Nazik olun: sitelere saniyede onlarca istek atmayin. Script her istek
  arasina time.sleep(1) koyuyor, bunu azaltmayin.

Kurulum:
    pip install requests beautifulsoup4 openpyxl pdfplumber --break-system-packages

Kullanim:
    python3 scrape_arms_sources.py --output turkiye_kaynak_taramasi.xlsx
    python3 scrape_arms_sources.py --output out.xlsx --debug
"""

import argparse
import re
import time
from dataclasses import dataclass, field

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept": (
        "text/html,application/xhtml+xml,application/xml;q=0.9,"
        "image/avif,image/webp,*/*;q=0.8"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
    "Sec-Fetch-Dest": "document",
    "Sec-Fetch-Mode": "navigate",
    "Sec-Fetch-Site": "none",
    "Sec-Fetch-User": "?1",
}

# Session kullaniyoruz ki cerezler (cookies) istekler arasinda korunsun --
# bazi siteler ilk istekte bir "gecici" cerez verip, ikinci istekte onu
# beklerler; duz requests.get() her seferinde sifirdan baglanir, Session
# bunu bir tarayici gibi surdurur.
SESSION = requests.Session()
SESSION.headers.update(HEADERS)

DSCA_TURKEY_URL = "https://www.dsca.mil/Press-Media/Major-Arms-Sales/Tag/58501/turkey"
STATE_DEPT_URL = "https://www.state.gov/arms-sales-congressional-notifications/"


@dataclass
class Notification:
    source: str
    date: str
    title: str
    url: str
    estimated_cost: str = ""
    summary: str = ""


def fetch(url: str, debug: bool = False) -> BeautifulSoup | None:
    """Ortak fetch fonksiyonu. Basarisiz olursa None doner."""
    try:
        resp = SESSION.get(url, timeout=15)
        resp.raise_for_status()
    except requests.RequestException as e:
        print(f"  [HATA] {url} cekilemedi: {e}")
        if debug and hasattr(e, "response") and e.response is not None:
            # 403 sayfasinin ilk 300 karakterini goster -- bu bize WAF'in
            # basit bir "Access Denied" sayfasi mi yoksa bir JS-challenge
            # (orn. "Just a moment...", "Checking your browser") sayfasi mi
            # dondurdugunu gosterir. Ikincisi ise Selenium/Playwright gerekir.
            snippet = e.response.text[:300].replace("\n", " ")
            print(f"  [DEBUG] Sunucu yaniti (ilk 300 karakter): {snippet}")
        return None

    if debug:
        print(f"  [DEBUG] {url} -> HTTP {resp.status_code}, {len(resp.text)} karakter")

    return BeautifulSoup(resp.text, "html.parser")


def scrape_dsca_turkey_selenium(debug: bool = False) -> list[Notification]:
    """
    DSCA'nin 'requests' kutuphanesini Akamai WAF uzerinden engellemesi
    (bkz. 'Access Denied, Reference #18...' hatasi) nedeniyle, bu fonksiyon
    GERCEK bir Chrome tarayicisini Selenium ile kontrol eder. Akamai artik
    gercek bir Chrome'un TLS/JS imzasini gorur, bu yuzden engellenmez.

    Kurulum: pip install selenium webdriver-manager
    (webdriver-manager, chromedriver'i sizin icin otomatik indirir --
    elle bir surucu indirmenize gerek yok.)
    """
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.common.by import By
    from webdriver_manager.chrome import ChromeDriverManager

    print("DSCA (Turkiye) taraniyor -- Selenium (gercek Chrome) ile...")

    options = Options()
    # NOT: headless (gorunmez) mod Akamai tarafindan ayri tespit edilebiliyor.
    # Once GORUNUR bir Chrome penceresiyle deneyin (asagidaki satir kapali).
    # Calisirsa ve ileride tekrar gizli calistirmak isterseniz, alttaki satiri
    # yorumdan cikarabilirsiniz -- ama once calistigindan emin olun.
    # options.add_argument("--headless=new")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--window-size=1280,900")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
    results = []
    try:
        driver.get(DSCA_TURKEY_URL)
        time.sleep(2)  # sayfanin tam yuklenmesini bekle

        blocks = driver.find_elements(By.CSS_SELECTOR, "div.info")
        if debug:
            print(f"  [DEBUG] {len(blocks)} 'div.info' blogu bulundu (Selenium)")
            if len(blocks) == 0:
                print(f"  [DEBUG] Sayfa basligi: {driver.title!r}")
                print(f"  [DEBUG] Sayfa kaynagi (ilk 500 karakter): {driver.page_source[:500]!r}")

        for block in blocks:
            try:
                date = block.find_element(By.CSS_SELECTOR, "p.date").text.strip()
            except Exception:
                date = ""
            try:
                link_el = block.find_element(By.CSS_SELECTOR, "p.title a")
                title = link_el.text.strip()
                url = link_el.get_attribute("href")
            except Exception:
                continue  # baslik/link yoksa bu blok islenemez
            try:
                summary = block.find_element(By.CSS_SELECTOR, "p.hidden-oxs").text.strip()
            except Exception:
                summary = ""

            results.append(Notification(source="DSCA", date=date, title=title, url=url, summary=summary))

        # Her bildirimin ayrinti sayfasina girip tam maliyet bilgisini cek
        for n in results:
            driver.get(n.url)
            time.sleep(1.5)
            body_text = driver.find_element(By.TAG_NAME, "body").text
            cost_match = re.search(r"estimated cost of \$([\d.,]+\s*(million|billion))", body_text, re.IGNORECASE)
            if cost_match:
                n.estimated_cost = cost_match.group(0)
    finally:
        driver.quit()

    print(f"  -> {len(results)} DSCA bildirimi bulundu (Selenium).")
    return results


def scrape_dsca_turkey(debug: bool = False) -> list[Notification]:
    """
    DSCA'nin Turkiye etiketli sayfasindan bildirimleri cek.

    NOT: Bu selector'lar TAHMIN DEGIL -- Chrome DevTools ile sayfa canli
    incelenerek dogrulandi. Gercek yapi:

        <div class="info">
            <p class="date">May 14, 2025</p>
            <p class="title"><a href="...">Baslik</a></p>
            <p class="hidden-oxs">WASHINGTON, ... ozet metni</p>
        </div>

    Site yapisini degistirirse (siteler zamanla degisir), ayni yontemi
    tekrarlayin: sayfayi Chrome'da acin, F12 (DevTools) -> Elements
    sekmesinde ilgili ogeye sag tik -> "Inspect" -> gercek class/etiket
    adlarini gorun, asagidaki select_one() cagrilarini ona gore guncelleyin.
    """
    print("DSCA (Turkiye) taraniyor...")
    soup = fetch(DSCA_TURKEY_URL, debug=debug)
    if soup is None:
        return []

    results = []
    info_blocks = soup.select("div.info")

    if debug:
        print(f"  [DEBUG] {len(info_blocks)} 'div.info' blogu bulundu")

    for block in info_blocks:
        date_tag = block.select_one("p.date")
        title_tag = block.select_one("p.title a")
        summary_tag = block.select_one("p.hidden-oxs")

        if not title_tag:
            continue

        url = title_tag.get("href", "")
        title = title_tag.get_text(strip=True)
        date = date_tag.get_text(strip=True) if date_tag else ""
        summary = summary_tag.get_text(strip=True) if summary_tag else ""

        results.append(Notification(source="DSCA", date=date, title=title, url=url, summary=summary))

    # Her bildirinin ayrinti sayfasina girip tam maliyet bilgisini cek
    # (liste sayfasindaki ozet cogu zaman kesik oluyor)
    for n in results:
        detail_soup = fetch(n.url, debug=debug)
        if detail_soup is None:
            continue
        text = detail_soup.get_text(" ", strip=True)
        cost_match = re.search(r"estimated cost of \$([\d.,]+\s*(million|billion))", text, re.IGNORECASE)
        if cost_match:
            n.estimated_cost = cost_match.group(0)
        time.sleep(1)  # siteye nazik davranin

    print(f"  -> {len(results)} DSCA bildirimi bulundu.")
    return results


def scrape_state_dept_turkey(debug: bool = False) -> list[Notification]:
    """
    State Department'in yeni (Subat 2026 sonrasi) bildirim sayfasindan
    Turkiye'ye dair kayitlari cek. Sayfa, ulke filtresi listesinde Turkiye
    gorunmuyorsa (yani hic bildirim yoksa) bos liste doner -- bu durumu
    ayrica raporlayin, cunku kendisi bir bulgu.
    """
    print("State Department (Turkiye) taraniyor...")
    soup = fetch(STATE_DEPT_URL, debug=debug)
    if soup is None:
        return []

    page_text = soup.get_text(" ", strip=True)
    if "Turkey" not in page_text and "Türkiye" not in page_text:
        print("  -> UYARI: Sayfada 'Turkey'/'Türkiye' hic gecmiyor. "
              "Subat 2026'dan bu yana yeni bir FMS bildirimi olmamis olabilir. "
              "Bu durumu notlarinizda mutlaka belirtin (dogrulanmasi gereken bir bulgu).")
        return []

    results = []
    links = soup.find_all("a", href=re.compile(r"/releases/bureau-of-political-military-affairs/"))
    for link in links:
        title = link.get_text(strip=True)
        if "turk" not in title.lower() and "türk" not in title.lower():
            continue
        url = link.get("href")
        if not url.startswith("http"):
            url = "https://www.state.gov" + url
        results.append(Notification(source="State Dept", date="", title=title, url=url))

    print(f"  -> {len(results)} State Dept bildirimi bulundu.")
    return results


def search_pdf_for_turkey(pdf_path: str, keyword: str = "Türkei") -> list[str]:
    """
    Indirdiginiz Alman Rüstungsexportbericht PDF'ini yerelde tarar,
    'keyword' (varsayilan: 'Türkei') gecen paragraflari dondurur.
    Once PDF'i elle indirip bu fonksiyona yolunu vermeniz gerekiyor --
    bkz. bundeswirtschaftsministerium.de/.../ruestungsexportbericht-*.pdf
    """
    import pdfplumber

    hits = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if keyword in text:
                # Anahtar kelime cevresindeki 300 karakteri al
                idx = text.find(keyword)
                snippet = text[max(0, idx - 150): idx + 150]
                hits.append(f"[Sayfa {i+1}] ...{snippet}...")
    return hits


def write_excel(dsca_results, state_results, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "DSCA_Bildirimleri"

    headers = ["Kaynak", "Tarih", "Baslik", "Tahmini_Maliyet", "URL", "Ozet"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    row = 2
    for n in dsca_results:
        ws.cell(row=row, column=1, value=n.source)
        ws.cell(row=row, column=2, value=n.date)
        ws.cell(row=row, column=3, value=n.title)
        ws.cell(row=row, column=4, value=n.estimated_cost)
        ws.cell(row=row, column=5, value=n.url)
        ws.cell(row=row, column=6, value=n.summary)
        row += 1

    ws2 = wb.create_sheet("State_Dept_Bildirimleri")
    for col, h in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
    row = 2
    for n in state_results:
        ws2.cell(row=row, column=1, value=n.source)
        ws2.cell(row=row, column=2, value=n.date)
        ws2.cell(row=row, column=3, value=n.title)
        ws2.cell(row=row, column=5, value=n.url)
        row += 1

    for sheet in (ws, ws2):
        for col_cells in sheet.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 60)

    wb.save(output_path)
    print(f"\nYazildi: {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="turkiye_kaynak_taramasi.xlsx")
    parser.add_argument("--debug", action="store_true")
    args = parser.parse_args()

    dsca_results = scrape_dsca_turkey(debug=args.debug)
    if not dsca_results:
        print("\n  'requests' ile DSCA'ya erisilemedi (Akamai engeli olabilir) "
              "-- Selenium (gercek Chrome) ile tekrar deneniyor...\n")
        try:
            dsca_results = scrape_dsca_turkey_selenium(debug=args.debug)
        except ImportError:
            print("  [HATA] Selenium kurulu degil. Kurmak icin: "
                  "pip3 install selenium webdriver-manager")
        except Exception as e:
            print(f"  [HATA] Selenium ile de basarisiz oldu: {e}")

    state_results = scrape_state_dept_turkey(debug=args.debug)
    write_excel(dsca_results, state_results, args.output)


if __name__ == "__main__":
    main()
