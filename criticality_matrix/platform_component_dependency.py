"""
platform_component_dependency.py
---------------------------------
SIPRI Arms Transfers Database (trade_register.csv) uzerinden, "yerli uretim"
(Local production = Yes) olarak isaretlenmis platformlarin, farkli bir
tedarikci ulkeden gelen kritik alt sistemlerle (Engines / Sensors / Naval
weapons / Other) zaman-ortakligi (temporal co-occurrence) uzerinden
eslestirilmesi.

ONEMLI METODOLOJIK UYARI
------------------------
SIPRI veri tabaninda platform ve alt sistem kayitlari arasinda ACIK bir
"bu motor bu ucaga takildi" baglantisi YOKTUR. Veri tabani sadece iki ayri
satir tutar (bkz. paylasilan metodoloji metni, * dipnotu). Bu script, ayni
alici ulkeye (varsayilan: Turkiye) yakin yillarda gelen platform ve alt
sistem kayitlarini ESLESTIRME ADAYI olarak isaretler. Bu bir HEURISTIC'tir,
KANITLANMIS bir tedarik zinciri baglantisi degildir. Gercek dogrulama icin
basin/ureticiw raporlari (Janes, Defense News, sirket web siteleri, SIPRI
Fact Sheet'leri vb.) ile capraz kontrol gereklidir. Script bu yuzden ayrica
manuel dogrulama icin bos bir "To_Verify_External" sablonu da uretir.

Kullanim:
    python3 platform_component_dependency.py \
        --input /path/to/trade_register.csv \
        --output /path/to/output.xlsx \
        --recipient Turkiye \
        --year-window 2
"""

import argparse
import csv
from collections import defaultdict

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Alt sistem / bilesen kategorileri (SIPRI metodolojisindeki * isaretli kategoriler)
SUBSYSTEM_CATEGORIES = {"Engines", "Sensors", "Naval weapons", "Other"}

# Kategori-bazli "makul eslesme" mantigi (tamamen kural-tabanli, bilgi
# amaclidir; SIPRI'nin kendisi bu baglantiyi kurmuyor)
PLAUSIBLE_LINKS = {
    "Aircraft": {"Engines", "Sensors", "Other"},
    "Ships": {"Engines", "Naval weapons", "Sensors", "Other"},
    "Armoured vehicles": {"Engines", "Sensors", "Other"},
    "Air-defence systems": {"Sensors", "Engines", "Other"},
    "Artillery": {"Sensors"},
    "Missiles": {"Sensors"},
    "Satellites": {"Sensors"},
}


def load_sipri_csv(path: str) -> pd.DataFrame:
    """SIPRI export dosyasindaki basliklar/ust bilgiler ve satir sonu ';'
    karakterlerini temizleyip duzgun bir DataFrame olarak yukler."""
    with open(path, encoding="utf-8-sig") as f:
        lines = f.readlines()

    header_idx = next(
        i for i, l in enumerate(lines) if l.startswith("SIPRI AT Database ID")
    )
    clean_lines = [l.rstrip("\r\n").rstrip(";") + "\n" for l in lines[header_idx:]]

    reader = csv.DictReader(clean_lines)
    rows = list(reader)
    df = pd.DataFrame(rows)

    # Sayisal alanlari donustur
    for col in ["Order date", "Delivery year", "Numbers delivered",
                "SIPRI estimate", "TIV deal unit", "TIV delivery values"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


def build_matches(df: pd.DataFrame, recipient: str, year_window: int) -> tuple:
    """Platform ve alt sistem tablolarini olusturur, aday eslesmeleri kurar."""

    recip_df = df[df["Recipient"] == recipient].copy()

    platforms = recip_df[
        (~recip_df["Armament category"].isin(SUBSYSTEM_CATEGORIES))
        & (recip_df["Local production"] == "Yes")
    ].copy()

    subsystems = recip_df[
        recip_df["Armament category"].isin(SUBSYSTEM_CATEGORIES)
    ].copy()

    matches = []
    for _, plat in platforms.iterrows():
        plat_year = plat["Delivery year"]
        if pd.isna(plat_year):
            continue
        plat_supplier = plat["Supplier"]
        plat_category = plat["Armament category"]

        candidates = subsystems[
            (subsystems["Supplier"] != plat_supplier)
            & (subsystems["Delivery year"].notna())
            & ((subsystems["Delivery year"] - plat_year).abs() <= year_window)
        ]

        for _, sub in candidates.iterrows():
            year_gap = int(sub["Delivery year"] - plat_year)
            plausible = sub["Armament category"] in PLAUSIBLE_LINKS.get(plat_category, set())
            matches.append({
                "Platform_SIPRI_ID": plat["SIPRI AT Database ID"],
                "Platform_Designation": plat["Designation"],
                "Platform_Description": plat["Description"],
                "Platform_Category": plat_category,
                "Platform_Supplier": plat_supplier,
                "Platform_Delivery_Year": int(plat_year),
                "Component_SIPRI_ID": sub["SIPRI AT Database ID"],
                "Component_Designation": sub["Designation"],
                "Component_Description": sub["Description"],
                "Component_Category": sub["Armament category"],
                "Component_Supplier": sub["Supplier"],
                "Component_Delivery_Year": int(sub["Delivery year"]),
                "Year_Gap": year_gap,
                "Plausible_Category_Link": "Evet" if plausible else "Belirsiz",
                "Match_Confidence": "Orta (kategori uyumlu)" if plausible else "Dusuk (sadece zaman ortakligi)",
            })

    matches_df = pd.DataFrame(matches)
    return platforms, subsystems, matches_df


def supplier_summary(subsystems: pd.DataFrame) -> pd.DataFrame:
    summary = (
        subsystems.groupby(["Supplier", "Armament category"])
        .agg(Deal_Count=("SIPRI AT Database ID", "count"),
             Total_TIV=("TIV delivery values", "sum"))
        .reset_index()
        .sort_values(["Armament category", "Total_TIV"], ascending=[True, False])
    )
    return summary


# ---------------------------------------------------------------------------
# Excel yazma
# ---------------------------------------------------------------------------

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
NOTE_FONT = Font(name="Arial", italic=True, size=10, color="555555")


def write_df_sheet(wb: Workbook, name: str, df: pd.DataFrame):
    ws = wb.create_sheet(name)
    if df.empty:
        ws["A1"] = "Bu kategori icin veri bulunamadi."
        ws["A1"].font = NOTE_FONT
        return ws

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    ws.freeze_panes = "A2"
    return ws


def write_readme(wb: Workbook, recipient: str, year_window: int,
                  n_platforms: int, n_subsystems: int, n_matches: int):
    ws = wb.create_sheet("README", 0)
    ws.column_dimensions["A"].width = 100

    lines = [
        ("Platform-Bilesen Bagimlilik Haritasi", TITLE_FONT),
        (f"Alici ulke: {recipient}  |  Eslestirme yil penceresi: +/- {year_window} yil", BODY_FONT),
        ("", BODY_FONT),
        ("YONTEM", Font(name="Arial", bold=True, size=11)),
        ("1. 'Platforms_LocalProduction' sekmesi: SIPRI'de 'Local production = Yes' "
         "olarak isaretlenmis, alicisi secilen ulke olan platform teslimatlari "
         "(Aircraft, Ships, Armoured vehicles, Artillery, Air-defence systems, "
         "Missiles, Satellites kategorileri).", BODY_FONT),
        ("2. 'Subsystem_Imports' sekmesi: Ayni ulkeye giden Engines, Sensors, "
         "Naval weapons, Other kategorisindeki TUM alt sistem ithalatlari.", BODY_FONT),
        ("3. 'Candidate_Matches' sekmesi: Bir platform teslimati ile, farkli bir "
         "tedarikci ulkeden gelen ve teslimat yili birbirine +/- "
         f"{year_window} yil icinde olan alt sistem teslimatlari 'aday eslesme' "
         "olarak esletirilir. 'Plausible_Category_Link' sutunu, kategori mantigina "
         "gore (orn. Aircraft + Engines) bu eslesmenin akla yatkin olup olmadigini "
         "isaretler.", BODY_FONT),
        ("4. 'Chain_Dependencies' sekmesi: Turkiye'nin SUPPLIER olarak (yani ihracatci "
         "olarak) gorundugu satirlari alip, ayni tasarim adiyla Turkiye'nin RECIPIENT "
         "oldugu (yani kendisinin bu sistemi hangi ulkeden, hangi kosulla aldigi) "
         "satirlarla eslestirir. Boylece 'Orijinal Tedarikci -> Turkiye (lisansli "
         "uretim) -> Ucuncu Ulke' seklinde uc katmanli bir zincir ortaya cikar. Bu, "
         "designation (tasarim adi) string eslesmesine dayanir; farkli varyant "
         "isimlendirmeleri (orn. 'AIFV' vs 'AIFV-APC') otomatik normallestirilir ama "
         "yine de manuel goz kontrolu onerilir.", BODY_FONT),
        ("", BODY_FONT),
        ("ONEMLI SINIRLAMALAR", Font(name="Arial", bold=True, size=11, color="C00000")),
        ("- Bu bir ISTATISTIKSEL/ZAMANSAL eslestirmedir, KANITLANMIS bir tedarik "
         "zinciri baglantisi DEGILDIR. SIPRI veri tabaninda hangi motorun hangi "
         "platforma takildigina dair acik bir referans alani yoktur.", NOTE_FONT),
        ("- SIPRI sadece 'major weapons' takip eder; hafif bilesenler (orn. bazi "
         "IHA motorlari/EO sensorleri, agirlik esiginin altinda kalan sistemler) "
         "bu veri tabaninda hic gorunmeyebilir. Bu tur bilinen vakalar icin "
         "'To_Verify_External' sekmesindeki sablonu kullanarak basin/endustri "
         "kaynaklarindan manuel dogrulama eklemeniz gerekir.", NOTE_FONT),
        ("- 'Local production: Yes' etiketi lisansli/yerli montaj anlamina "
         "gelebilir; bu, bilesen duzeyinde tam teknolojik bagimsizlik anlamina "
         "gelmez.", NOTE_FONT),
        ("- Duplike SIPRI ID'ler (ayni antlasmanin farkli yillara yayilan "
         "teslimatlari) birden fazla satirda gorunebilir; bu kasitlidir, cunku "
         "her satir o yila ait fiili teslimati temsil eder.", NOTE_FONT),
        ("", BODY_FONT),
        ("OZET SAYILAR", Font(name="Arial", bold=True, size=11)),
        (f"- Yerli-uretim platform teslimati: {n_platforms}", BODY_FONT),
        (f"- Alt sistem (Engines/Sensors/Naval weapons/Other) ithalati: {n_subsystems}", BODY_FONT),
        (f"- Aday eslesme (platform x bilesen): {n_matches}", BODY_FONT),
        ("", BODY_FONT),
        ("Uretim tarihi: bu script calistirildiginda otomatik olusur. "
         "Kaynak: SIPRI Arms Transfers Database (trade_register.csv).", NOTE_FONT),
    ]

    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return ws


def write_to_verify_template(wb: Workbook):
    cols = [
        "Platform_Designation", "Component_Type", "Claimed_Component_Supplier_Country",
        "Source_Name", "Source_URL", "Source_Date", "Verification_Status",
        "Notes",
    ]
    example_row = [
        "Bayraktar TB-2", "EO/IR sensor", "Kanada (orn. Wescam - eski donem)",
        "orn. Janes / Defense News haberi", "https://...", "2020-XX-XX",
        "Dogrulanmadi / Dogrulandi / Celiskili",
        "SIPRI veri tabaninda agirlik esigi altinda kaldigi icin gorunmuyor",
    ]
    ws = wb.create_sheet("To_Verify_External")
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for col_idx, val in enumerate(example_row, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(name="Arial", italic=True, size=10, color="808080")
    for col_idx in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
    ws.freeze_panes = "A2"
    return ws


# ---------------------------------------------------------------------------
# EK: Uc katmanli zincirleme bagimlilik analizi
# (Orijinal Tedarikci -> Turkiye [lisansli uretim] -> Ucuncu Ulke)
# ---------------------------------------------------------------------------
import re as _re

def build_chain_dependencies(df: pd.DataFrame) -> pd.DataFrame:
    def norm(name):
        n = name.strip()
        n = _re.sub(r'\s*-+\s*$', '', n)
        n = _re.sub(r'\s*\(.*?\)\s*', '', n)
        return n.strip()

    df = df.copy()
    df['Designation_norm'] = df['Designation'].apply(norm)

    tr_exports = df[df['Supplier'] == 'Turkiye']
    tr_own_source = df[(df['Recipient'] == 'Turkiye') & (df['Supplier'] != 'Turkiye')]

    chains = []
    for design_norm in tr_exports['Designation_norm'].unique():
        export_rows = tr_exports[tr_exports['Designation_norm'] == design_norm]
        source_rows = tr_own_source[tr_own_source['Designation_norm'] == design_norm]
        if source_rows.empty:
            continue
        origin_suppliers = source_rows['Supplier'].unique().tolist()
        origin_local_prod = source_rows['Local production'].unique().tolist()
        third_countries = export_rows['Recipient'].unique().tolist()
        chains.append({
            'Designation': design_norm,
            'Description': export_rows['Description'].iloc[0],
            'Category': export_rows['Armament category'].iloc[0],
            'Original_Supplier_to_Turkiye': ', '.join(origin_suppliers),
            'Turkiye_Local_Production': ', '.join(origin_local_prod),
            'Turkiye_Export_Recipients': ', '.join(third_countries),
            'N_Third_Countries': len(third_countries),
            'Total_Export_TIV': round(export_rows['TIV delivery values'].sum(), 1),
        })
    return pd.DataFrame(chains).sort_values('N_Third_Countries', ascending=False)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default="/mnt/user-data/uploads/trade_register.csv")
    parser.add_argument("--output", default="/mnt/user-data/outputs/turkiye_platform_bilesen_bagimliligi.xlsx")
    parser.add_argument("--recipient", default="Turkiye")
    parser.add_argument("--year-window", type=int, default=2)
    args = parser.parse_args()

    df = load_sipri_csv(args.input)
    platforms, subsystems, matches_df = build_matches(df, args.recipient, args.year_window)
    summary_df = supplier_summary(subsystems)

    display_platform_cols = [
        "SIPRI AT Database ID", "Designation", "Description", "Armament category",
        "Supplier", "Order date", "Delivery year", "Numbers delivered", "Status",
        "TIV delivery values", "Local production",
    ]
    display_subsystem_cols = display_platform_cols  # ayni sema

    wb = Workbook()
    wb.remove(wb.active)  # varsayilan bos sayfayi kaldir

    write_readme(wb, args.recipient, args.year_window,
                 len(platforms), len(subsystems), len(matches_df))
    write_df_sheet(wb, "Platforms_LocalProduction", platforms[display_platform_cols].sort_values("Delivery year"))
    write_df_sheet(wb, "Subsystem_Imports", subsystems[display_subsystem_cols].sort_values("Delivery year"))
    write_df_sheet(wb, "Candidate_Matches",
                    matches_df.sort_values(["Platform_Delivery_Year", "Match_Confidence"]) if not matches_df.empty else matches_df)
    write_df_sheet(wb, "Supplier_Summary", summary_df)

    chains_df = build_chain_dependencies(df)
    write_df_sheet(wb, "Chain_Dependencies", chains_df)

    write_to_verify_template(wb)

    wb.save(args.output)
    print(f"Yazildi: {args.output}")
    print(f"  Platformlar: {len(platforms)}")
    print(f"  Alt sistem ithalatlari: {len(subsystems)}")
    print(f"  Aday eslesmeler: {len(matches_df)}")


if __name__ == "__main__":
    main()


